"""
Enrollment Prediction Report Generator
-------------------------------------------------------------------------------
Generates comprehensive Excel reports (XLSX) for enrollment forecasting.

Workbook contents:
1) Overview & Synopsis
   - Report metadata (timestamp, model info)
   - Natural-language synopsis
   - Optional key performance metrics (from per-course accuracy CSV)

2) Visualizations
   - Accuracy distribution pie chart (if accuracy_df provided)
   - Top 10 most predictable courses bar chart
   - MAPE distribution histogram-style bar chart

3) Enrollment Predictions
   - Hierarchical table:
     Institution → School → Department → Course rows
   - Includes historical enrollment (from DB) for recent years
   - Includes forecast enrollment for the target year

4) Accuracy Analysis (optional)
   - Table view of per-course accuracy with conditional formatting on MAPE

Inputs:
- predictions_json: list-of-dicts returned by /api/predict/sql (recommended)
- OR predictions_csv: path to CSV containing predictions in expected columns
- Optional accuracy_csv: per-course accuracy metrics CSV

Notes / Design choices:
- Historical enrollment is fetched live from Postgres (if reachable).
- Table discovery is dynamic (searches information_schema for required columns).
- Styling is done with openpyxl to keep the workbook readable and “report-like”.

Security note:
- This script reads from the DB using credentials from env vars. Ensure those are
  properly set in production environments.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
import psycopg2
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from psycopg2 import sql


# ==============================================================================
# Configuration
# ==============================================================================

DB_CONFIG = {
    # Mirrors backend/app/database/db_init.py defaults so we hit the same database
    "host": os.getenv("DB_HOST", "db"),
    "port": int(os.getenv("DB_PORT", 5432)),
    "user": os.getenv("POSTGRES_USER", "postgres"),
    "password": os.getenv("POSTGRES_PASSWORD", ""),
    "database": os.getenv("POSTGRES_DB", "postgres"),
}


@dataclass(frozen=True)
class ModelMetadata:
    """
    Small container for model metadata shown in the report.

    This is *display-only* information; it does not influence predictions.
    """
    model_type: str = "Machine Learning"
    feature_schema: str = "Standard"
    model_file: str = "N/A"


# ==============================================================================
# Data loading
# ==============================================================================

def load_predictions_from_json(json_data: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Convert JSON prediction data from the API endpoint to a DataFrame.

    Expected JSON format (examples):
    [
      {"subj": "MIS", "crse": 460, "term": 202540, "prediction": 76.98, "act": 20, "credits": 3, ...},
      ...
    ]

    Returns:
        DataFrame normalized to column names used by the report generator:
        - Subject
        - Course
        - Term
        - Credits
        - Predicted_Enrollment
        - Actual_Enrollment (optional, if present)
    """
    print(f"Loading {len(json_data)} predictions from JSON data...")
    df = pd.DataFrame(json_data)

    # Map API keys to report-friendly names
    column_mapping = {
        "subj": "Subject",
        "crse": "Course",
        "term": "Term",
        "credits": "Credits",
        "prediction": "Predicted_Enrollment",
        "act": "Actual_Enrollment",  # optional
    }
    df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})

    # Make predicted enrollment integer for presentation (forecast counts)
    if "Predicted_Enrollment" in df.columns:
        df["Predicted_Enrollment"] = df["Predicted_Enrollment"].round().astype(int)

    if "Subject" in df.columns and "Predicted_Enrollment" in df.columns:
        print(f"Loaded predictions for {df['Subject'].nunique()} unique subjects")
        print(f"Total predicted enrollment: {df['Predicted_Enrollment'].sum()}")
        print(f"Average predicted enrollment: {df['Predicted_Enrollment'].mean():.1f}")

    return df


# ==============================================================================
# Report generator
# ==============================================================================

class EnrollmentReportGenerator:
    """
    Generates the Excel workbook.

    Usage:
        generator = EnrollmentReportGenerator(predictions_df, accuracy_df, model_info)
        generator.generate_report("path/to/report.xlsx")
    """

    def __init__(
        self,
        predictions_df: pd.DataFrame,
        accuracy_df: Optional[pd.DataFrame] = None,
        model_info: Optional[Dict[str, str]] = None,
    ):
        """
        Args:
            predictions_df: DataFrame containing predictions (from JSON or CSV).
            accuracy_df: Optional per-course accuracy table with at least 'MAPE' column.
            model_info: Optional dict with keys: model_type, feature_schema, model_file.
        """
        self.predictions_df = predictions_df
        self.accuracy_df = accuracy_df

        # Normalize model metadata into a consistent structure
        self.model_metadata = ModelMetadata(
            model_type=(model_info or {}).get("model_type", "Machine Learning"),
            feature_schema=(model_info or {}).get("feature_schema", "Standard"),
            model_file=(model_info or {}).get("model_file", "N/A"),
        )

        # --- Excel styling presets (kept here so all sheets use consistent styling) ---
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        self.title_font = Font(name="Calibri", size=16, bold=True, color="366092")
        self.subtitle_font = Font(name="Calibri", size=12, bold=True, color="44546A")

        self.thin_side = Side(style="thin")
        self.border = Border(
            left=self.thin_side,
            right=self.thin_side,
            top=self.thin_side,
            bottom=self.thin_side,
        )

        # Table styling for hierarchical predictions sheet
        self.table_header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.table_subheader_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.course_identifier_fill = self.table_header_fill
        self.historical_data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.forecast_data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Cache discovered enrollment table name so we only discover once per run
        self._enrollment_table_name: Optional[str] = None

    # --------------------------------------------------------------------------
    # Public API
    # --------------------------------------------------------------------------

    def generate_report(self, output_path: str) -> None:
        """
        Build and write the full workbook to disk.

        Args:
            output_path: Where to save the .xlsx file
        """
        print("Generating enrollment prediction report...")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        print("  Creating Overview sheet...")
        self._create_overview_sheet(wb)

        print("  Creating Visualizations sheet...")
        self._create_visualizations_sheet(wb)

        print("  Creating Predictions sheet...")
        self._create_predictions_sheet(wb)

        if self.accuracy_df is not None:
            print("  Creating Accuracy Analysis sheet...")
            self._create_accuracy_sheet(wb)

        wb.save(output_path)
        print(f"✓ Report saved to: {output_path}")

    # --------------------------------------------------------------------------
    # Sheet: Overview & Synopsis
    # --------------------------------------------------------------------------

    def _create_overview_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create the “Overview & Synopsis” sheet with report context and summary."""
        ws = wb.create_sheet("Overview & Synopsis")

        # Title
        ws["A1"] = "Enrollment Prediction Report"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:D1")

        # Metadata block
        ws["A3"] = "Report Generated:"
        ws["B3"] = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        ws["A4"] = "Model Type:"
        ws["B4"] = self.model_metadata.model_type.title()
        ws["A5"] = "Feature Schema:"
        ws["B5"] = self.model_metadata.feature_schema.upper()
        ws["A6"] = "Model File:"
        ws["B6"] = self.model_metadata.model_file

        for r in range(3, 7):
            ws[f"A{r}"].font = Font(bold=True)

        # Synopsis section
        ws["A8"] = "Model Synopsis"
        ws["A8"].font = self.subtitle_font
        ws.merge_cells("A8:D8")

        synopsis_text = self._generate_synopsis()
        row = 10
        for line in synopsis_text.split("\n"):
            ws[f"A{row}"] = line
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        # Optional key metrics section
        if self.accuracy_df is not None:
            ws[f"A{row + 1}"] = "Key Performance Metrics"
            ws[f"A{row + 1}"].font = self.subtitle_font
            ws.merge_cells(f"A{row + 1}:D{row + 1}")

            metrics_row = row + 3
            self._add_key_metrics(ws, metrics_row)

        # Prediction summary section
        summary_row = row + 10
        ws[f"A{summary_row}"] = "Prediction Summary"
        ws[f"A{summary_row}"].font = self.subtitle_font
        ws.merge_cells(f"A{summary_row}:D{summary_row}")

        summary_data = self._get_prediction_summary()
        summary_row += 2
        for key, value in summary_data.items():
            ws[f"A{summary_row}"] = key
            ws[f"B{summary_row}"] = value
            ws[f"A{summary_row}"].font = Font(bold=True)
            summary_row += 1

        # Column widths for readability
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

    def _generate_synopsis(self) -> str:
        """
        Produce narrative text explaining the report, model, and usage guidance.

        This is deliberately human-readable and safe to include in a report packet.
        """
        model_type = self.model_metadata.model_type

        synopsis = (
            f"This report presents enrollment predictions generated using a {model_type} model. "
            f"The predictions are based on historical enrollment data including course identifiers, "
            f"term information, section details, and credit hours.\n\n"
        )

        if self.accuracy_df is not None and "MAPE" in self.accuracy_df.columns:
            total_courses = len(self.accuracy_df)
            good_courses = (self.accuracy_df["MAPE"] < 20).sum()
            pct_good = (good_courses / total_courses * 100) if total_courses else 0.0

            synopsis += (
                "Model Performance: "
                f"The model demonstrates excellent predictive accuracy for {good_courses} out of "
                f"{total_courses} courses ({pct_good:.1f}%), with a Mean Absolute Percentage Error "
                "(MAPE) below 20%. These courses represent stable, predictable enrollment patterns "
                "suitable for operational planning.\n\n"
            )

        synopsis += (
            "Usage Recommendations: Predictions should be used as a planning tool alongside "
            "institutional knowledge. Courses with lower prediction counts or unusual patterns "
            "should be reviewed manually. The model is most accurate for core required courses "
            "with consistent historical enrollment."
        )

        return synopsis

    def _add_key_metrics(self, ws, start_row: int) -> None:
        """
        Add key performance metrics derived from `accuracy_df` to the Overview sheet.
        """
        if self.accuracy_df is None or "MAPE" not in self.accuracy_df.columns:
            return

        total_courses = len(self.accuracy_df)
        excellent = (self.accuracy_df["MAPE"] < 20).sum()
        good = ((self.accuracy_df["MAPE"] >= 20) & (self.accuracy_df["MAPE"] < 40)).sum()
        poor = (self.accuracy_df["MAPE"] >= 40).sum()
        avg_mape = self.accuracy_df["MAPE"].mean()
        median_mape = self.accuracy_df["MAPE"].median()

        # Header row
        ws[f"A{start_row}"] = "Metric"
        ws[f"B{start_row}"] = "Value"
        for c in ("A", "B"):
            ws[f"{c}{start_row}"].font = self.header_font
            ws[f"{c}{start_row}"].fill = self.header_fill

        # Metric rows
        metrics = [
            ("Total Courses Analyzed", total_courses),
            ("Excellent Predictions (MAPE < 20%)", f"{excellent} ({excellent/total_courses*100:.1f}%)"),
            ("Good Predictions (MAPE 20-40%)", f"{good} ({good/total_courses*100:.1f}%)"),
            ("Poor Predictions (MAPE > 40%)", f"{poor} ({poor/total_courses*100:.1f}%)"),
            ("Average MAPE", f"{avg_mape:.2f}%"),
            ("Median MAPE", f"{median_mape:.2f}%"),
        ]

        for i, (metric, value) in enumerate(metrics, start=1):
            ws[f"A{start_row + i}"] = metric
            ws[f"B{start_row + i}"] = value
            ws[f"A{start_row + i}"].border = self.border
            ws[f"B{start_row + i}"].border = self.border

    def _get_prediction_summary(self) -> Dict[str, Any]:
        """Compute basic summary stats from the predictions dataframe."""
        if "Predicted_Enrollment" not in self.predictions_df.columns:
            return {
                "Total Predictions": len(self.predictions_df),
                "Unique Courses": "N/A",
                "Predicted Total Enrollment": "N/A",
                "Average Predicted Enrollment": "N/A",
            }

        unique_courses = "N/A"
        if {"Subject", "Course"}.issubset(self.predictions_df.columns):
            unique_courses = self.predictions_df[["Subject", "Course"]].drop_duplicates().shape[0]

        return {
            "Total Predictions": len(self.predictions_df),
            "Unique Courses": unique_courses,
            "Predicted Total Enrollment": int(self.predictions_df["Predicted_Enrollment"].sum()),
            "Average Predicted Enrollment": f"{self.predictions_df['Predicted_Enrollment'].mean():.1f}",
        }

    # --------------------------------------------------------------------------
    # Helpers: Term / course normalization & history lookup
    # --------------------------------------------------------------------------

    def _determine_forecast_year(self) -> int:
        """
        Infer forecast year from Term codes (YYYYxx), defaulting to next calendar year.
        """
        if "Term" in self.predictions_df.columns:
            years: List[int] = []
            for term_value in self.predictions_df["Term"].dropna():
                year = self._extract_year_from_term(term_value)
                if year:
                    years.append(year)
            if years:
                return max(years)
        return datetime.now().year + 1

    def _get_historical_years(self, forecast_year: int, count: int = 3) -> List[int]:
        """
        Return a list of historical years to display (e.g., last 3 years).
        """
        years = [forecast_year - count + i for i in range(count)]
        return [y for y in years if y > 0]

    @staticmethod
    def _normalize_subject_code(subject: Any) -> Optional[str]:
        """Normalize subject codes (e.g., ' fin ' → 'FIN')."""
        if pd.isna(subject):
            return None
        s = str(subject).strip().upper()
        return s or None

    @staticmethod
    def _normalize_course_number(course: Any) -> Optional[str]:
        """
        Normalize course identifiers into a consistent string.
        - If numeric, drop trailing .0
        - Otherwise keep the trimmed string
        """
        if pd.isna(course):
            return None
        try:
            numeric = float(course)
            if numeric.is_integer():
                return str(int(numeric))
        except (TypeError, ValueError):
            pass
        s = str(course).strip()
        return s or None

    @staticmethod
    def _course_number_sort_key(course_value: Any) -> float:
        """
        Sorting helper for course numbers:
        attempts to extract digits and sort numerically.
        """
        normalized = EnrollmentReportGenerator._normalize_course_number(course_value)
        if not normalized:
            return float("inf")
        digits = "".join(ch for ch in str(normalized) if ch.isdigit())
        return int(digits) if digits else float("inf")

    @staticmethod
    def _extract_year_from_term(term_value: Any) -> Optional[int]:
        """
        Extract year from term code (e.g., 202540 → 2025).
        """
        try:
            term_int = int(float(term_value))
            return term_int // 100 if term_int > 0 else None
        except (TypeError, ValueError):
            return None

    def _get_db_connection(self):
        """
        Connect to Postgres for historical enrollment lookups.

        Returns:
            psycopg2 connection or None if connection fails (report still generates).
        """
        try:
            return psycopg2.connect(**DB_CONFIG)
        except Exception as exc:
            print(f"[WARN] Unable to connect to PostgreSQL for historical data: {exc}")
            return None

    def _discover_enrollment_table(self, conn) -> Optional[str]:
        """
        Identify a table containing subj/crse/term/act columns (needed for history).
        Returns the first match found in the public schema.
        """
        discovery_sql = """
            SELECT table_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND column_name IN ('subj', 'crse', 'term', 'act')
            GROUP BY table_name
            HAVING COUNT(DISTINCT column_name) = 4
            ORDER BY table_name
            LIMIT 1;
        """
        with conn.cursor() as cur:
            cur.execute(discovery_sql)
            row = cur.fetchone()
        return row[0] if row else None

    def _fetch_historical_enrollment(
        self,
        subjects: List[str],
        courses: List[str],
        years: List[int],
    ) -> Dict[Tuple[str, str, int], int]:
        """
        Fetch summed historical enrollment per (subject, course, year) from Postgres.

        Returns:
            dict keyed by (SUBJ, CRSE, YEAR) → total enrollment (int)
        """
        if not subjects or not courses or not years:
            return {}

        conn = self._get_db_connection()
        if not conn:
            return {}

        try:
            if not self._enrollment_table_name:
                self._enrollment_table_name = self._discover_enrollment_table(conn)
                if not self._enrollment_table_name:
                    print("[WARN] Could not locate enrollment table with subj/crse/term/act columns.")
                    return {}

            history_sql = sql.SQL(
                """
                SELECT subj,
                       crse,
                       FLOOR(NULLIF(term::text, '')::numeric / 100)::INT AS hist_year,
                       SUM(COALESCE(act, 0)) AS total_enrollment
                FROM {table}
                WHERE subj = ANY(%s)
                  AND crse::text = ANY(%s)
                  AND NULLIF(term::text, '') IS NOT NULL
                  AND act IS NOT NULL
                  AND FLOOR(NULLIF(term::text, '')::numeric / 100)::INT = ANY(%s)
                GROUP BY subj, crse, hist_year
                """
            ).format(table=sql.Identifier(self._enrollment_table_name))

            with conn.cursor() as cur:
                cur.execute(history_sql, (subjects, courses, years))
                rows = cur.fetchall()

            history_map: Dict[Tuple[str, str, int], int] = {}
            for subj, crse, year, total in rows:
                norm_subj = self._normalize_subject_code(subj)
                norm_course = self._normalize_course_number(crse)
                if norm_subj and norm_course and year is not None:
                    history_map[(norm_subj, norm_course, int(year))] = int(round(total or 0))

            return history_map

        except Exception as exc:
            print(f"[WARN] Failed to fetch historical enrollments: {exc}")
            return {}
        finally:
            conn.close()

    def _set_cell_border(self, cell, left=False, right=False, top=False, bottom=False) -> None:
        """
        Apply borders selectively to a cell (useful for merged/structured tables).
        """
        def _side(flag: bool) -> Side:
            return self.thin_side if flag else Side(border_style=None)

        cell.border = Border(
            left=_side(left),
            right=_side(right),
            top=_side(top),
            bottom=_side(bottom),
        )

    # --------------------------------------------------------------------------
    # Sheet: Visualizations
    # --------------------------------------------------------------------------

    def _create_visualizations_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create the “Visualizations” sheet (charts depend on accuracy_df)."""
        ws = wb.create_sheet("Visualizations")

        ws["A1"] = "Enrollment Prediction Visualizations"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:F1")

        if self.accuracy_df is None:
            ws["A3"] = "No accuracy data provided. Charts are unavailable."
            return

        # Charts
        self._add_accuracy_pie_chart(ws, start_row=3)
        self._add_top_courses_chart(ws, start_row=20)
        self._add_mape_distribution(ws, start_row=37)

    def _add_accuracy_pie_chart(self, ws, start_row: int) -> None:
        """Pie chart: Excellent / Good / Poor accuracy distribution by MAPE bins."""
        ws[f"A{start_row}"] = "Model Accuracy Distribution"
        ws[f"A{start_row}"].font = self.subtitle_font

        excellent = (self.accuracy_df["MAPE"] < 20).sum()
        good = ((self.accuracy_df["MAPE"] >= 20) & (self.accuracy_df["MAPE"] < 40)).sum()
        poor = (self.accuracy_df["MAPE"] >= 40).sum()

        data_row = start_row + 2
        ws[f"A{data_row}"] = "Category"
        ws[f"B{data_row}"] = "Count"
        ws[f"A{data_row}"].font = Font(bold=True)
        ws[f"B{data_row}"].font = Font(bold=True)

        ws[f"A{data_row + 1}"] = "Excellent (< 20%)"
        ws[f"B{data_row + 1}"] = excellent
        ws[f"A{data_row + 2}"] = "Good (20-40%)"
        ws[f"B{data_row + 2}"] = good
        ws[f"A{data_row + 3}"] = "Poor (> 40%)"
        ws[f"B{data_row + 3}"] = poor

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_row + 3)
        data = Reference(ws, min_col=2, min_row=data_row, max_row=data_row + 3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Prediction Accuracy by Category"
        pie.height = 10
        pie.width = 15

        ws.add_chart(pie, f"D{start_row}")

    def _add_top_courses_chart(self, ws, start_row: int) -> None:
        """Bar chart: top 10 lowest MAPE courses."""
        ws[f"A{start_row}"] = "Top 10 Most Predictable Courses"
        ws[f"A{start_row}"].font = self.subtitle_font

        top_10 = self.accuracy_df.nsmallest(10, "MAPE")

        data_row = start_row + 2
        ws[f"A{data_row}"] = "Course"
        ws[f"B{data_row}"] = "MAPE %"
        ws[f"A{data_row}"].font = Font(bold=True)
        ws[f"B{data_row}"].font = Font(bold=True)

        for i, (_, r) in enumerate(top_10.iterrows(), start=1):
            course_name = f"{r['Subject']} {r['Course']}"
            ws[f"A{data_row + i}"] = course_name
            ws[f"B{data_row + i}"] = round(r["MAPE"], 2)

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top 10 Most Predictable Courses (Lowest MAPE)"
        chart.y_axis.title = "MAPE %"
        chart.x_axis.title = "Course"

        data = Reference(ws, min_col=2, min_row=data_row, max_row=data_row + 10)
        cats = Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_row + 10)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 15

        ws.add_chart(chart, f"D{start_row}")

    def _add_mape_distribution(self, ws, start_row: int) -> None:
        """Bar chart: histogram-style distribution of MAPE ranges."""
        ws[f"A{start_row}"] = "MAPE Distribution Across All Courses"
        ws[f"A{start_row}"].font = self.subtitle_font

        bins = [0, 10, 20, 30, 40, 50, 100, 500, 3000]
        labels = ["0-10%", "10-20%", "20-30%", "30-40%", "40-50%", "50-100%", "100-500%", "500%+"]

        mape_values = self.accuracy_df["MAPE"].dropna()
        counts, _ = np.histogram(mape_values, bins=bins)

        data_row = start_row + 2
        ws[f"A{data_row}"] = "MAPE Range"
        ws[f"B{data_row}"] = "Course Count"
        ws[f"A{data_row}"].font = Font(bold=True)
        ws[f"B{data_row}"].font = Font(bold=True)

        for i, (label, count) in enumerate(zip(labels, counts), start=1):
            ws[f"A{data_row + i}"] = label
            ws[f"B{data_row + i}"] = int(count)

        chart = BarChart()
        chart.type = "col"
        chart.style = 11
        chart.title = "Distribution of Prediction Accuracy"
        chart.y_axis.title = "Number of Courses"
        chart.x_axis.title = "MAPE Range"

        data = Reference(ws, min_col=2, min_row=data_row, max_row=data_row + len(labels))
        cats = Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_row + len(labels))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 15

        ws.add_chart(chart, f"D{start_row}")

    # --------------------------------------------------------------------------
    # Sheet: Enrollment Predictions
    # --------------------------------------------------------------------------

    def _create_predictions_sheet(self, wb: openpyxl.Workbook) -> None:
        """
        Create the “Enrollment Predictions” sheet with a hierarchical table that includes:
        - Department headers
        - Course rows
        - Historical enrollment (from DB, if available)
        - Forecast enrollment (from predictions)
        """
        ws = wb.create_sheet("Enrollment Predictions")

        # Top header / metadata
        ws["A1"] = "Enrollment Prediction Report"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:G1")

        row = 2
        ws[f"A{row}"] = "Report Generated:"
        ws[f"B{row}"] = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        row += 1

        ws[f"A{row}"] = "Version:"
        ws[f"B{row}"] = "0.01"
        row += 2

        # Model card
        ws[f"A{row}"] = "Model Card"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        ws[f"A{row}"] = "Model Type:"
        ws[f"B{row}"] = self.model_metadata.model_type.title()
        row += 1

        ws[f"A{row}"] = "Feature Schema:"
        ws[f"B{row}"] = self.model_metadata.feature_schema.upper()
        row += 1

        ws[f"A{row}"] = "Model File:"
        ws[f"B{row}"] = self.model_metadata.model_file
        row += 2

        # Query summary
        ws[f"A{row}"] = "Prediction Query"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        ws[f"A{row}"] = "Semester(s):"
        if "Term" in self.predictions_df.columns and len(self.predictions_df) > 0:
            terms = sorted(self.predictions_df["Term"].dropna().unique())
            ws[f"B{row}"] = ", ".join([str(t) for t in terms])
        row += 1

        ws[f"A{row}"] = "Filter:"
        row += 1

        ws[f"A{row}"] = "Prediction Summary"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        ws[f"A{row}"] = "Total Predictions"
        ws[f"B{row}"] = len(self.predictions_df)
        row += 1

        ws[f"A{row}"] = "MAPE"
        if self.accuracy_df is not None and "MAPE" in self.accuracy_df.columns:
            ws[f"B{row}"] = f"{self.accuracy_df['MAPE'].mean():.0f}"
        row += 2

        # Section heading
        ws[f"A{row}"] = "Predictions"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        # Institution / school headers
        institution_row = row
        ws[f"A{institution_row}"] = "Central Connecticut State University"
        ws[f"A{institution_row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{institution_row}"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws.merge_cells(f"A{institution_row}:E{institution_row}")
        row += 1

        school_row = row
        ws[f"A{school_row}"] = "School of Business"
        ws[f"A{school_row}"].font = Font(bold=True, size=10, italic=True, color="FFFFFF")
        ws[f"A{school_row}"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws.merge_cells(f"A{school_row}:E{school_row}")
        row += 1

        forecast_year = self._determine_forecast_year()
        historical_years = self._get_historical_years(forecast_year)

        # Build prediction lookup: (SUBJ, CRSE) → predicted enrollment (int)
        pred_map: Dict[Tuple[str, str], int] = {}
        normalized_pairs: List[Tuple[str, str]] = []

        if {"Subject", "Course"}.issubset(self.predictions_df.columns):
            for _, r in self.predictions_df.iterrows():
                subj = self._normalize_subject_code(r.get("Subject"))
                course = self._normalize_course_number(r.get("Course"))
                prediction_value = r.get("Predicted_Enrollment", r.get("prediction"))
                if subj and course and pd.notna(prediction_value):
                    pred_map[(subj, course)] = int(round(float(prediction_value)))
                    normalized_pairs.append((subj, course))

        unique_subjects = sorted({subj for subj, _ in normalized_pairs})
        unique_courses = sorted({course for _, course in normalized_pairs})

        # Pull historical enrollment from DB (best-effort)
        history_map = self._fetch_historical_enrollment(unique_subjects, unique_courses, historical_years)

        # Group predictions by subject
        if {"Subject", "Course"}.issubset(self.predictions_df.columns):
            grouped = self.predictions_df.groupby("Subject")

            for subject, group_df in grouped:
                # Department header row
                dept_row = row
                dept_name = self._get_department_name(subject)

                ws[f"A{dept_row}"] = dept_name
                ws[f"A{dept_row}"].font = Font(bold=True, size=11)
                ws.merge_cells(f"A{dept_row}:E{dept_row}")

                dept_fill = PatternFill(start_color="BFD9EF", end_color="BFD9EF", fill_type="solid")
                for col in ["A", "B", "C", "D", "E"]:
                    cell = ws[f"{col}{dept_row}"]
                    cell.fill = dept_fill
                    self._set_cell_border(cell, left=(col == "A"), right=(col == "E"), top=True, bottom=True)

                row += 1

                # Table header row
                header_row = row
                ws[f"A{header_row}"] = "Course Identifier"
                ws[f"A{header_row}"].font = Font(bold=True)
                ws[f"A{header_row}"].fill = self.table_header_fill
                self._set_cell_border(ws[f"A{header_row}"], left=True, right=True, top=True, bottom=True)

                ws[f"B{header_row}"] = "Historical Enrollment"
                ws[f"B{header_row}"].font = Font(bold=True)
                ws.merge_cells(f"B{header_row}:D{header_row}")
                for col in ["B", "C", "D"]:
                    ws[f"{col}{header_row}"].fill = self.table_header_fill

                ws[f"E{header_row}"] = "Forecast"
                ws[f"E{header_row}"].font = Font(bold=True)
                ws[f"E{header_row}"].fill = self.table_header_fill

                ws[f"B{header_row}"].alignment = Alignment(horizontal="center")
                ws[f"E{header_row}"].alignment = Alignment(horizontal="center")

                self._set_cell_border(ws[f"B{header_row}"], left=True, top=True, bottom=True)
                self._set_cell_border(ws[f"C{header_row}"], top=True, bottom=True)
                self._set_cell_border(ws[f"D{header_row}"], right=True, top=True, bottom=True)
                self._set_cell_border(ws[f"E{header_row}"], left=True, right=True, top=True, bottom=True)

                row += 1

                # Subheader row: years
                sub_row = row
                ws[f"A{sub_row}"] = ""
                ws[f"A{sub_row}"].fill = self.course_identifier_fill
                self._set_cell_border(ws[f"A{sub_row}"], left=True, right=True, top=True, bottom=True)

                # Historical years occupy columns B-D (assumes 3 historical years)
                for i, year in enumerate(historical_years, start=2):
                    col_letter = get_column_letter(i)
                    ws[f"{col_letter}{sub_row}"] = str(year)
                    ws[f"{col_letter}{sub_row}"].font = Font(bold=True)
                    ws[f"{col_letter}{sub_row}"].alignment = Alignment(horizontal="center")
                    ws[f"{col_letter}{sub_row}"].fill = self.table_subheader_fill
                    self._set_cell_border(
                        ws[f"{col_letter}{sub_row}"],
                        left=(col_letter == "B"),
                        right=(col_letter == "D"),
                        top=True,
                        bottom=True,
                    )

                ws[f"E{sub_row}"] = str(forecast_year)
                ws[f"E{sub_row}"].font = Font(bold=True)
                ws[f"E{sub_row}"].alignment = Alignment(horizontal="center")
                ws[f"E{sub_row}"].fill = self.forecast_data_fill
                self._set_cell_border(ws[f"E{sub_row}"], left=True, right=True, top=True, bottom=True)
                row += 1

                # Course list for this department, sorted
                courses = group_df[["Course", "Subject"]].drop_duplicates()
                courses = (
                    courses.assign(_course_sort=courses["Course"].apply(self._course_number_sort_key))
                    .sort_values(by=["_course_sort", "Course"])
                    .drop(columns="_course_sort")
                )

                # Course rows
                for _, course_row_data in courses.iterrows():
                    subj_code = self._normalize_subject_code(course_row_data["Subject"])
                    course_code = self._normalize_course_number(course_row_data["Course"])

                    ws[f"A{row}"] = f"{course_row_data['Subject']} {course_row_data['Course']}"
                    ws[f"A{row}"].alignment = Alignment(horizontal="left")
                    ws[f"A{row}"].fill = self.course_identifier_fill
                    self._set_cell_border(ws[f"A{row}"], left=True, right=True, top=True, bottom=True)

                    # Historical values
                    for i, year in enumerate(historical_years, start=2):
                        col_letter = get_column_letter(i)
                        value = ""
                        if subj_code and course_code:
                            value = history_map.get((subj_code, course_code, year), "")
                        ws[f"{col_letter}{row}"] = value
                        ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center")
                        ws[f"{col_letter}{row}"].fill = self.historical_data_fill
                        self._set_cell_border(
                            ws[f"{col_letter}{row}"],
                            left=(col_letter == "B"),
                            right=(col_letter == "D"),
                            top=True,
                            bottom=True,
                        )

                    # Forecast
                    pred_value = pred_map.get((subj_code, course_code), "") if subj_code and course_code else ""
                    ws[f"E{row}"] = pred_value
                    ws[f"E{row}"].alignment = Alignment(horizontal="center")
                    ws[f"E{row}"].fill = self.forecast_data_fill
                    self._set_cell_border(ws[f"E{row}"], left=True, right=True, top=True, bottom=True)

                    row += 1

                # Spacer row between departments
                row += 1

        # Column widths
        ws.column_dimensions["A"].width = 20
        uniform_year_width = 12
        ws.column_dimensions["B"].width = uniform_year_width
        ws.column_dimensions["C"].width = uniform_year_width
        ws.column_dimensions["D"].width = uniform_year_width
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12

    def _get_department_name(self, subject_code: str) -> str:
        """
        Map subject codes to department names.

        NOTE:
        - Consider replacing this hardcoded mapping with your subjectDepartmentMap.json
          for a single source of truth across frontend/backend/reporting.
        """
        dept_mapping = {
            "AC": "Accounting",
            "FIN": "Finance",
            "MGT": "Management",
            "MIS": "Management Information Systems",
            "MKT": "Marketing",
            "BUS": "Business Administration",
            "LAW": "Business Law & Ethics",
            "MC": "Managerial Communication",
        }
        return dept_mapping.get(subject_code, f"{subject_code} Department")

    # --------------------------------------------------------------------------
    # Sheet: Accuracy Analysis (optional)
    # --------------------------------------------------------------------------

    def _create_accuracy_sheet(self, wb: openpyxl.Workbook) -> None:
        """
        Create a sheet with the per-course accuracy table and conditional formatting on MAPE.
        """
        ws = wb.create_sheet("Accuracy Analysis")

        ws["A1"] = "Per-Course Accuracy Analysis"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:H1")

        start_row = 3
        for r_idx, row in enumerate(
            dataframe_to_rows(self.accuracy_df, index=False, header=True),
            start=start_row,
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Header styling
                if r_idx == start_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # MAPE conditional colors
                    if "MAPE" in self.accuracy_df.columns:
                        mape_col_idx = self.accuracy_df.columns.get_loc("MAPE") + 1
                        if c_idx == mape_col_idx and isinstance(value, (int, float)) and not np.isnan(value):
                            if value < 20:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif value < 40:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                cell.border = self.border

        # Auto-fit columns (bounded to avoid absurd widths)
        for col_idx in range(1, len(self.accuracy_df.columns) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(start_row, start_row + len(self.accuracy_df) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        # Add filters
        last_col_letter = ws.cell(row=start_row, column=len(self.accuracy_df.columns)).column_letter
        ws.auto_filter.ref = f"A{start_row}:{last_col_letter}{start_row + len(self.accuracy_df)}"


# ==============================================================================
# Orchestration function
# ==============================================================================

def generate_enrollment_report(
    predictions_json: Optional[List[Dict[str, Any]]] = None,
    predictions_csv: Optional[str] = None,
    accuracy_csv: Optional[str] = None,
    output_path: Optional[str] = None,
    model_info: Optional[Dict[str, str]] = None,
) -> str:
    """
    Generate an enrollment prediction Excel report.

    Args:
        predictions_json: Predictions returned from API (list of dicts).
        predictions_csv: Path to CSV file with predictions (fallback).
        accuracy_csv: Path to per-course accuracy CSV (optional).
        output_path: Where to save the workbook (auto-generated if omitted).
        model_info: Dict containing report metadata (model_type, feature_schema, model_file).

    Returns:
        The output path where the report was saved.
    """
    # --- Load predictions ---
    if predictions_json is not None:
        print("Loading predictions from JSON data...")
        predictions_df = load_predictions_from_json(predictions_json)

        # Optional convenience: also persist a CSV snapshot
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        predictions_csv_path = Path(__file__).parent.parent / "reports" / f"predictions_{timestamp}.csv"
        predictions_csv_path.parent.mkdir(parents=True, exist_ok=True)
        predictions_df.to_csv(predictions_csv_path, index=False)
        print(f"Predictions saved to: {predictions_csv_path}")

    elif predictions_csv is not None:
        print(f"Loading predictions from {predictions_csv}...")
        predictions_df = pd.read_csv(predictions_csv)
    else:
        raise ValueError("Either predictions_json or predictions_csv must be provided")

    # --- Load accuracy data (optional) ---
    accuracy_df = pd.read_csv(accuracy_csv) if accuracy_csv else None

    # --- Determine output path ---
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = Path(__file__).parent.parent / "reports"
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(out_dir / f"enrollment_report_{timestamp}.xlsx")

    # --- Generate report ---
    print("\nGenerating Excel report...")
    generator = EnrollmentReportGenerator(predictions_df, accuracy_df, model_info)
    generator.generate_report(output_path)

    # Friendly output (container + Windows path hint)
    print("\n" + "=" * 80)
    print("✓ Report saved successfully!")
    print("=" * 80)
    print(f"Container path: {output_path}")

    if output_path.startswith("/app/"):
        windows_path = (
            output_path.replace("/app/", "C:\\Development\\ccsu_enrollment\\Enrollment-Predictor\\")
            .replace("/", "\\")
        )
        print(f"Windows path:   {windows_path}")

    return output_path


# ==============================================================================
# CLI entrypoint
# ==============================================================================

def _build_arg_parser() -> argparse.ArgumentParser:
    """Build the CLI parser (kept separate for readability/testability)."""
    parser = argparse.ArgumentParser(
        description="Generate enrollment prediction Excel reports from API data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate report with predictions from JSON file (from API endpoint)
  python generate_report.py --json predictions.json

  # Generate report with existing predictions CSV
  python generate_report.py --predictions predictions.csv

  # Generate report with accuracy data
  python generate_report.py --json predictions.json --accuracy backend/app/ml/test_results/per_course_accuracy_tree_20251109_013953.csv

  # With custom model info
  python generate_report.py --json predictions.json --model-type "Tree Ensemble" --model-name "enrollment_tree_v2"
        """,
    )

    parser.add_argument("--json", help="Path to JSON file with predictions from API endpoint")
    parser.add_argument("--predictions", help="Path to CSV file with predictions")
    parser.add_argument("--accuracy", help="Path to CSV file with per-course accuracy metrics")
    parser.add_argument("--output", help="Output path for Excel report (auto-generated if not provided)")
    parser.add_argument("--model-type", help="Model type for report metadata (e.g., 'Tree Ensemble', 'Neural Network')")
    parser.add_argument("--model-name", help="Model name/file for report metadata")
    parser.add_argument("--feature-schema", help="Feature schema used (e.g., 'min', 'rich', 'auto')")

    return parser


if __name__ == "__main__":
    parser = _build_arg_parser()
    args = parser.parse_args()

    if args.json is None and args.predictions is None:
        parser.error("Either --json or --predictions must be provided")

    try:
        # Load JSON predictions if requested
        predictions_json = None
        if args.json:
            print(f"Loading JSON data from {args.json}...")
            with open(args.json, "r", encoding="utf-8") as f:
                predictions_json = json.load(f)

        # Build model metadata if any fields provided
        model_info = None
        if args.model_type or args.model_name or args.feature_schema:
            model_info = {
                "model_type": args.model_type or "Machine Learning",
                "feature_schema": args.feature_schema or "Standard",
                "model_file": args.model_name or "N/A",
            }

        output = generate_enrollment_report(
            predictions_json=predictions_json,
            predictions_csv=args.predictions,
            accuracy_csv=args.accuracy,
            output_path=args.output,
            model_info=model_info,
        )

        print(f"\n✓ Report generation complete: {output}")

    except Exception as e:
        print(f"\n✗ Error generating report: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
